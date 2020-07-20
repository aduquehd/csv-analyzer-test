from openpyxl import load_workbook
from datetime import timedelta
from datetime import date
from datetime import datetime


def analyze_file(dataset_file):
    """
    Analyze a dataset xlsx file, getting all the data into a list of dictionaries.
    :param dataset_file: DataSetFiles objects.
    :return: List of xlsx file columns data.
    """
    data = []
    data_set_id = str(dataset_file.data_set.id)
    data_set_file_id = str(dataset_file.id)

    wb = load_workbook(filename=dataset_file.file, read_only=True)
    ws = wb['daily_weather']

    object_date = datetime.combine(dataset_file.start_date, datetime.min.time())

    for row in ws.iter_rows(min_row=2):
        row_data = {
            'data_set_id': data_set_id,
            'data_set_file_id': data_set_file_id,
            'date': object_date,
        }
        object_date = object_date + timedelta(days=1)
        for cell in row:
            try:
                row_data[ws[f'{cell.column_letter}1'].value] = cell.value
            except AttributeError:
                continue
        data.append(row_data)

    return data
